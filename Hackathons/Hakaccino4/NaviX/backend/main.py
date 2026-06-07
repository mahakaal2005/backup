import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import uuid
from datetime import datetime
from typing import List, Optional
import io
import openpyxl
import xlrd

from database import get_db, init_db
from models import (
    ChatRequest, ChatResponse, Session, ThinkingLogSummary,
    Report, HeatmapEntry, EquipmentItem
)
from config import REPORTS_DIR

app = FastAPI(title="Navix API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if os.path.exists(REPORTS_DIR):
    app.mount("/reports", StaticFiles(directory=REPORTS_DIR), name="reports")

@app.on_event("startup")
def startup():
    init_db()

@app.get("/")
def root():
    return {"status": "ok", "service": "Navix API", "version": "2.0.0"}

@app.get("/health")
def health():
    from agents.base_agent import _llm_cache
    return {"status": "healthy", "cache_entries": len(_llm_cache)}

@app.post("/api/clear-cache")
def clear_cache():
    from agents.base_agent import clear_llm_cache
    clear_llm_cache()
    return {"status": "ok", "message": "LLM response cache cleared."}

@app.delete("/api/sessions/{session_id}")
def delete_session(session_id: str):
    conn = get_db()
    try:
        conn.execute("DELETE FROM thoughts WHERE conversation_id IN (SELECT conversation_id FROM conversations WHERE session_id = ?)", (session_id,))
        conn.execute("DELETE FROM conversations WHERE session_id = ?", (session_id,))
        conn.execute("DELETE FROM reports WHERE session_id = ?", (session_id,))
        conn.execute("DELETE FROM sessions WHERE session_id = ?", (session_id,))
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.post("/api/chat")
async def chat(request: ChatRequest):
    from agents.orchestrator import process_chat
    result = await process_chat(request.message, request.session_id)
    return result

@app.get("/api/session-ids")
def get_session_ids():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT session_id, user_query, session_date FROM sessions ORDER BY created_at DESC"
        ).fetchall()
        return [
            {
                "session_id": row["session_id"],
                "user_query": row["user_query"] or "",
                "session_date": row["session_date"] or "",
            }
            for row in rows
        ]
    finally:
        conn.close()

@app.get("/api/sessions/{session_id}")
def get_session(session_id: str):
    conn = get_db()
    try:
        session_row = conn.execute(
            "SELECT * FROM sessions WHERE session_id = ?", (session_id,)
        ).fetchone()

        if not session_row:
            raise HTTPException(status_code=404, detail="Session not found")

        conversations = conn.execute(
            "SELECT * FROM conversations WHERE session_id = ? ORDER BY created_at ASC",
            (session_id,)
        ).fetchall()

        conv_list = []
        for conv in conversations:
            thoughts = conn.execute(
                "SELECT * FROM thoughts WHERE conversation_id = ? ORDER BY thought_id ASC",
                (conv["conversation_id"],)
            ).fetchall()
            messages = [
                {
                    "user_query": conv["user_query"] or "",
                    "agent_output": conv["agent_response"] or "",
                    "agent_name": "Chatbot",
                    "action": "",
                }
            ]
            conv_list.append({
                "conversation_id": conv["conversation_id"],
                "user_query": conv["user_query"] or "",
                "agent_response": conv["agent_response"] or "",
                "messages": messages,
                "thoughts": [
                    {
                        "agent_name": t["agent_name"],
                        "thinking_stage": t["thinking_stage"],
                        "thought_content": t["thought_content"],
                        "output_content": t["output_content"],
                        "created_date": t["created_date"],
                    }
                    for t in thoughts
                ],
            })

        return {
            "session_id": session_id,
            "user_query": session_row["user_query"] or "",
            "session_date": session_row["session_date"] or "",
            "conversations": conv_list,
        }
    finally:
        conn.close()

@app.get("/api/reports")
def get_reports():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT report_id, session_id, blob_url, created_at FROM reports ORDER BY created_at DESC"
        ).fetchall()
        return [
            {
                "report_id": row["report_id"],
                "session_id": row["session_id"],
                "blob_url": row["blob_url"],
                "created_at": row["created_at"],
            }
            for row in rows
        ]
    finally:
        conn.close()

@app.get("/api/reports/download/{report_id}")
def download_report(report_id: str):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT file_path, blob_url FROM reports WHERE report_id = ?",
            (report_id,)
        ).fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Report not found")

        file_path = row["file_path"]
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Report file not found on disk")

        filename = os.path.basename(file_path)
        return FileResponse(
            path=file_path,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename=filename,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    finally:
        conn.close()

@app.get("/api/reports/by-session/{session_id}")
def get_reports_by_session(session_id: str):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT report_id, session_id, blob_url, file_path, created_at FROM reports WHERE session_id = ? ORDER BY created_at DESC",
            (session_id,)
        ).fetchall()
        return [
            {
                "report_id": row["report_id"],
                "session_id": row["session_id"],
                "blob_url": row["blob_url"],
                "download_url": f"/api/reports/download/{row['report_id']}",
                "created_at": row["created_at"],
            }
            for row in rows
        ]
    finally:
        conn.close()

@app.get("/api/thinking-log-ids")
def get_thinking_log_ids():
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT s.session_id, s.user_query
            FROM sessions s
            WHERE EXISTS (
                SELECT 1 FROM thoughts t
                JOIN conversations c ON t.conversation_id = c.conversation_id
                WHERE c.session_id = s.session_id
            )
            ORDER BY s.created_at DESC
            """
        ).fetchall()
        return [
            {
                "id": row["session_id"],
                "session_id": row["session_id"],
                "first_query": row["user_query"] or "Untitled session",
            }
            for row in rows
        ]
    finally:
        conn.close()

@app.get("/api/thinking-logs-by-session-id/{session_id}")
def get_thinking_logs_by_session(session_id: str):
    conn = get_db()
    try:
        session_row = conn.execute(
            "SELECT * FROM sessions WHERE session_id = ?", (session_id,)
        ).fetchone()

        if not session_row:
            raise HTTPException(status_code=404, detail="Session not found")

        conversations = conn.execute(
            "SELECT * FROM conversations WHERE session_id = ? ORDER BY created_at ASC",
            (session_id,)
        ).fetchall()

        conv_list = []
        for conv in conversations:
            raw_thoughts = conn.execute(
                "SELECT * FROM thoughts WHERE conversation_id = ? ORDER BY thought_id ASC",
                (conv["conversation_id"],)
            ).fetchall()

            agent_map = {}
            for t in raw_thoughts:
                agent = t["agent_name"]
                if agent not in agent_map:
                    agent_map[agent] = []
                agent_map[agent].append({
                    "thought_content": t["thought_content"],
                    "thinking_stage": t["thinking_stage"],
                    "output_content": t["output_content"],
                    "thinking_stage_output": t["output_content"],
                    "created_date": t["created_date"],
                })

            agents_list = [
                {"agent_name": agent, "thoughts": thoughts}
                for agent, thoughts in agent_map.items()
            ]

            conv_list.append({
                "conversation_id": conv["conversation_id"],
                "user_query": conv["user_query"] or "",
                "agents": agents_list,
            })

        return {
            "session_id": session_id,
            "conversations": conv_list,
        }
    finally:
        conn.close()

@app.get("/api/heatmap")
def get_heatmap(conversation_id: Optional[str] = None, session_id: Optional[str] = None):
    conn = get_db()
    try:
        country_rows = conn.execute(
            """
            SELECT vendor_country as country, AVG(risk_level) as avg_risk
            FROM equipment_schedule
            GROUP BY vendor_country
            ORDER BY avg_risk DESC
            """
        ).fetchall()

        item_rows = conn.execute(
            """
            SELECT item_name, vendor_country, status, risk_level,
                   expected_date, actual_date, project_country
            FROM equipment_schedule
            ORDER BY risk_level DESC
            """
        ).fetchall()

        from agents.scheduler_agent import _calculate_variance, _categorize_risk

        items_by_country: dict = {}
        for r in item_rows:
            c = r["vendor_country"]
            if c not in items_by_country:
                items_by_country[c] = []
            variance = _calculate_variance(r["expected_date"], r["actual_date"])
            category = _categorize_risk(dict(r))
            items_by_country[c].append({
                "item_name": r["item_name"],
                "status": r["status"],
                "risk_level": r["risk_level"],
                "risk_category": category,
                "variance_days": variance,
                "expected_date": r["expected_date"] or "",
                "actual_date": r["actual_date"] or "",
                "project_country": r["project_country"] or "",
            })

        result = []
        for row in country_rows:
            avg = float(row["avg_risk"])
            if avg >= 4.5:
                risk_label = "5"
            elif avg >= 3.5:
                risk_label = "4"
            elif avg >= 2.5:
                risk_label = "3"
            elif avg >= 1.5:
                risk_label = "2"
            else:
                risk_label = "1"

            result.append({
                "country": row["country"],
                "average_risk": risk_label,
                "items": items_by_country.get(row["country"], []),
            })

        return result
    finally:
        conn.close()

@app.get("/api/equipment")
def get_equipment():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM equipment_schedule ORDER BY id ASC").fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()

REQUIRED_COLUMNS = {"item name", "expected date", "actual date", "vendor country", "project country"}


def _parse_date_value(val, field: str, row_num: int):
    """Parse a date value from an Excel cell into a YYYY-MM-DD string.

    Handles: Python datetime/date objects (openpyxl), xlrd float serial dates
    (converted upstream), and string representations in YYYY-MM-DD or MM/DD/YYYY.
    Returns (date_str_or_None, error_str_or_None).
    """
    from datetime import date as date_type
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d"), None
    if isinstance(val, date_type):
        return val.strftime("%Y-%m-%d"), None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d"), None
        except ValueError:
            pass
    return None, f"Row {row_num}: Could not parse {field} date '{val}'."


def _read_rows_from_xlsx(contents: bytes):
    """Parse an .xlsx file and return a list of tuples (one per row, including header)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Could not parse the .xlsx file. Make sure it is a valid Excel file.") from exc
    ws = wb.active
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _read_rows_from_xls(contents: bytes):
    """Parse a legacy .xls file and return a list of tuples (one per row, including header).

    xlrd represents dates as floats; we convert them to date strings using xlrd's
    xldate_as_datetime helper.
    """
    try:
        wb = xlrd.open_workbook(file_contents=contents)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Could not parse the .xls file. Make sure it is a valid Excel file.") from exc
    ws = wb.sheet_by_index(0)
    rows = []
    for row_idx in range(ws.nrows):
        cells = []
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt_val = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    cells.append(dt_val)
                except Exception:
                    cells.append(cell.value)
            else:
                cells.append(cell.value)
        rows.append(tuple(cells))
    return rows


def _build_items_from_rows(raw_rows: list):
    """Validate headers and data rows; return a list of insertion tuples."""
    if not raw_rows:
        raise HTTPException(status_code=400, detail="The file has no rows.")

    header_row = raw_rows[0]
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(sorted(missing))}. "
                   f"Expected columns: item name, expected date, actual date, vendor country, project country."
        )

    col_idx = {col: headers.index(col) for col in REQUIRED_COLUMNS}

    items = []
    parse_errors = []
    for row_num, row in enumerate(raw_rows[1:], start=2):
        if len(row) <= max(col_idx.values()):
            parse_errors.append(f"Row {row_num}: too few columns.")
            continue

        item_name = row[col_idx["item name"]]
        vendor_country = row[col_idx["vendor country"]]
        project_country = row[col_idx["project country"]]
        expected_date_raw = row[col_idx["expected date"]]
        actual_date_raw = row[col_idx["actual date"]]

        if not item_name and not vendor_country and not project_country:
            continue

        if not item_name or not vendor_country or not project_country:
            parse_errors.append(f"Row {row_num}: item name, vendor country, and project country are required.")
            continue

        expected_date, err1 = _parse_date_value(expected_date_raw, "expected", row_num)
        actual_date, err2 = _parse_date_value(actual_date_raw, "actual", row_num)

        if err1:
            parse_errors.append(err1)
            continue
        if err2:
            parse_errors.append(err2)
            continue

        status = "on_track"
        risk_level = 1
        if expected_date and actual_date:
            try:
                exp = datetime.strptime(expected_date, "%Y-%m-%d").date()
                act = datetime.strptime(actual_date, "%Y-%m-%d").date()
                variance = (act - exp).days
                if variance > 20:
                    status = "delayed"
                    risk_level = 4
                elif variance > 7:
                    status = "delayed"
                    risk_level = 3
                elif variance > 0:
                    status = "at_risk"
                    risk_level = 2
            except Exception:
                pass

        items.append((str(item_name).strip(), str(vendor_country).strip(), str(project_country).strip(),
                      expected_date, actual_date, status, risk_level))

    if parse_errors:
        raise HTTPException(
            status_code=422,
            detail="File contains errors:\n" + "\n".join(parse_errors)
        )

    if not items:
        raise HTTPException(status_code=400, detail="No valid data rows found in the file.")

    return items


@app.post("/api/upload-schedule")
async def upload_schedule(file: UploadFile = File(...)):
    filename = (file.filename or "").lower()
    is_xlsx = filename.endswith(".xlsx")
    is_xls = filename.endswith(".xls") and not is_xlsx

    if not is_xlsx and not is_xls:
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted.")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")

    if is_xlsx:
        raw_rows = _read_rows_from_xlsx(contents)
    else:
        raw_rows = _read_rows_from_xls(contents)

    items = _build_items_from_rows(raw_rows)

    conn = get_db()
    try:
        conn.execute("DELETE FROM equipment_schedule")
        conn.executemany(
            """INSERT INTO equipment_schedule
               (item_name, vendor_country, project_country, expected_date, actual_date, status, risk_level)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            items
        )
        conn.commit()
    finally:
        conn.close()

    return {"status": "success", "items_loaded": len(items)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
